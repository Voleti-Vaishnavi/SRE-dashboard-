import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Application, Team
from app.utils.enums import ChangeStatus, FourEyeStatus, RunStatus, UrlStatus
from app.utils.excel_utils import (
    add_dropdown,
    autosize_columns,
    style_example_row,
    style_header_row,
)

YESTERDAY = date.today() - timedelta(days=1)


def _app_names(db: Session) -> list[str]:
    return [name for (name,) in db.query(Application.name).order_by(Application.name).all()]


def _team_names(db: Session) -> list[str]:
    return [name for (name,) in db.query(Team.name).order_by(Team.name).all()]


def _build_workbook(
    headers: list[str],
    example_row: list,
    dropdowns: dict[str, list[str]],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    ws.append(example_row)

    style_header_row(ws, len(headers))
    style_example_row(ws, len(headers))
    autosize_columns(ws, headers)
    ws.freeze_panes = "A2"
    ws["A2"].comment = Comment(
        "EXAMPLE ROW — delete this row or overwrite it before uploading.",
        "SRE Dashboard",
    )

    lists_ws = wb.create_sheet("Lists")
    for i, (col_letter, values) in enumerate(dropdowns.items(), start=1):
        add_dropdown(ws, lists_ws, i, values, col_letter)
    lists_ws.sheet_state = "hidden"

    return wb


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_change_deployments_template(db: Session) -> bytes:
    headers = [
        "Application Name",
        "Tower Name",
        "CR Number",
        "Assignment Group",
        "Description",
        "Assigned To",
        "Change Status",
        "4-Eye Review Status",
        "Date",
    ]
    example = [
        "RAIS Internet",
        "RCIS",
        "CHG0001234",
        "RCIS Deployment Team",
        "Deploy latest release to production",
        "Jane Doe",
        ChangeStatus.SUCCESS.value,
        FourEyeStatus.COMPLETED.value,
        YESTERDAY,
    ]
    dropdowns = {
        "A": _app_names(db),
        "B": _team_names(db),
        "G": [s.value for s in ChangeStatus],
        "H": [s.value for s in FourEyeStatus],
    }
    wb = _build_workbook(headers, example, dropdowns)
    return _to_bytes(wb)


def build_url_availability_template(db: Session) -> bytes:
    headers = ["Application Name", "Date", "Status"]
    example = ["RAIS Internet", YESTERDAY, UrlStatus.OPERATIONAL.value]
    dropdowns = {
        "A": _app_names(db),
        "C": [s.value for s in UrlStatus],
    }
    wb = _build_workbook(headers, example, dropdowns)
    return _to_bytes(wb)


def _named_item_template(db: Session, item_label: str, example_item_name: str) -> bytes:
    headers = ["Application Name", item_label, "Date", "Status"]
    example = ["RAIS Internet", example_item_name, YESTERDAY, RunStatus.SUCCESS.value]
    dropdowns = {
        "A": _app_names(db),
        "D": [s.value for s in RunStatus],
    }
    wb = _build_workbook(headers, example, dropdowns)
    return _to_bytes(wb)


def build_job_monitoring_template(db: Session) -> bytes:
    return _named_item_template(db, "Job Name", "Nightly Batch Job")


def build_interface_monitoring_template(db: Session) -> bytes:
    return _named_item_template(db, "Interface Name", "Payment Gateway Interface")


def build_critical_file_monitoring_template(db: Session) -> bytes:
    return _named_item_template(db, "File Name", "EOD_Reconciliation.csv")


TEMPLATE_BUILDERS = {
    "change-deployments": (build_change_deployments_template, "change_deployments_template.xlsx"),
    "url-availability": (build_url_availability_template, "url_availability_template.xlsx"),
    "job-monitoring": (build_job_monitoring_template, "job_monitoring_template.xlsx"),
    "interface-monitoring": (
        build_interface_monitoring_template,
        "interface_monitoring_template.xlsx",
    ),
    "critical-file-monitoring": (
        build_critical_file_monitoring_template,
        "critical_file_monitoring_template.xlsx",
    ),
}
