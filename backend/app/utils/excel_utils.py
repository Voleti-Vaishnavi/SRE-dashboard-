from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def add_dropdown(
    ws: Worksheet,
    lists_ws: Worksheet,
    lists_col_idx: int,
    values: list[str],
    target_col_letter: str,
    max_row: int = 500,
) -> None:
    """Write `values` to a column on the hidden lists sheet and apply a dropdown
    data validation referencing that range to `target_col_letter` on `ws`."""
    list_col_letter = get_column_letter(lists_col_idx)
    lists_ws.cell(row=1, column=lists_col_idx, value=target_col_letter)
    for i, value in enumerate(values, start=2):
        lists_ws.cell(row=i, column=lists_col_idx, value=value)

    formula = f"Lists!${list_col_letter}$2:${list_col_letter}${len(values) + 1}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = "Please choose a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{target_col_letter}2:{target_col_letter}{max_row}")


def style_header_row(ws: Worksheet, num_columns: int) -> None:
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font


def style_example_row(ws: Worksheet, num_columns: int, row: int = 2) -> None:
    from openpyxl.styles import Font

    italic_gray = Font(italic=True, color="808080")
    for col in range(1, num_columns + 1):
        ws.cell(row=row, column=col).font = italic_gray


def autosize_columns(ws: Worksheet, headers: list[str], min_width: int = 14) -> None:
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, len(header) + 4)
